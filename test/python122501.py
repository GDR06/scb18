# -*- coding: utf-8 -*- 
# @Time : 2020/12/25 13:47 
# @Author : GDR 
# @File : python122501.py

import openpyxl

"""
接口自动化测试：
1、excel测试用例准备ok，代码可以自动读取用例数据
2、执行接口测试，得到响应结果                api_fun(url,data)
3、响应结果和预期结果比较（判断）----断言---通过/不通过  
4、最终执行结果通过与否，写入--excel表格

第三方库：openpyxl    1)安装  pip/pycharm   2)导入 import

Excel 中的三大对象
工作簿：WorkBook
表单：Sheet
单元格：Cell
"""

# cell = sheet.cell(row = 6,column = 7).value      # 获取第一行第一列单元格的值

# 读取excel数据---封装成函数
def case_read(filename,sheetname):           # 将读取excel数据的代码封装成函数
    wb = openpyxl.load_workbook(filename)    # 加载工作簿，打开一个excel文件-->test_case_api.xlsx
    sheet = wb[sheetname]                    # 打开一个表单
    max_row = sheet.max_row                  # 获取最大行数
    case_list = []                          # 新建空列表，存放for循环依次读取到的测试用例数据
    for i in range(2,max_row + 1):
        data_dict = dict(                   # 把每一行读取到的测试用例数据生成字典
        case_id = sheet.cell(row = i,column = 1).value,
        url = sheet.cell(row = i,column = 5).value,                    # 读取URL的值
        data = sheet.cell(row = i,column = 6).value,                     # 读取data的值
        expect = sheet.cell(row = i,column = 7).value)                 # 读取期望
        case_list.append(data_dict)                                   # 把生成的字典存放到List中

    return case_list

# 调用函数
case_data = case_read('../test_data/test_case_api.xlsx', 'login')
print(case_data)

# 写入数据到excel中
# 1、写入单元格的数据
# 2、保存，保存之前要关闭excel文档

def write_case(filename,sheetname,row,column,result):
    wb = openpyxl.load_workbook(filename)               # 加载工作簿，打开一个excel文件-->test_case_api.xlsx
    sheet = wb[sheetname]                               # 打开一个表单
    max_row = sheet.max_row
    for i in range(row,max_row + 1):
        sheet.cell(row = i,column = column).value = result
        wb.save(filename)

# 调用函数
write_case('../test_data/test_case_api.xlsx', 'register', 2, 8, 'Passed')
