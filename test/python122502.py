# -*- coding: utf-8 -*- 
# @Time : 2020/12/25 18:10 
# @Author : GDR 
# @File : python122502.py


"""
接口自动化测试：
1、excel测试用例准备ok，代码可以自动读取用例数据
2、执行接口测试，得到响应结果                api_fun(url,data)
3、响应结果和预期结果比较（判断）----断言---通过/不通过
4、最终执行结果通过与否，写入--excel表格

"""
import openpyxl
import requests

# 1、读取excel文件数据，封装成函数
def read_case(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    row_max = sheet.max_row
    list1 = []
    for i in range(2,row_max + 1):
        case_dict = dict(
            case_id = sheet.cell(row = i,column = 1).value,
            url = sheet.cell(row = i,column = 5).value,
            data = sheet.cell(row = i,column = 6).value,
            expect = sheet.cell(row = i,column = 7).value
        )
        list1.append(case_dict)
    return list1

# 2、执行接口测试(发送请求)，得到响应结果
def api_fun(url,data):
    headers = {'X-Lemonban-Media-Type': 'lemonban.v2', 'Content-Type': 'application/json'}
    res = requests.post(url = url,json = data,headers = headers).json()
    return res

# 3、断言：实际结果 == 预期结果
# 封装成函数
def execute_fun(filename,sheetname):
    case_list = read_case(filename,sheetname)    # 读数据，定义变量接收读取到的数据
    # eval()函数：运行被字符串包裹着的表达式
    for m in case_list:                                          # 遍历一个列表，每一个列表元素都是一条字典形式的数据
        case_id = m['case_id']                                   # 获取第几条用例
        response_dict = api_fun(url = m['url'],data = eval(m['data']))
        # real_code = response_dict['code']
        # expect_code = eval(m['expect'])['code']
        real_msg = response_dict['msg']
        expect_msg = eval(m['expect'])['msg']
        print("期望结果为：{}".format(expect_msg))
        print("实际结果为：{}".format(real_msg))
        if real_msg == expect_msg:
            print('第{}条用例执行通过！'.format(case_id))
            final_re = 'Passed'
        else:
            print('第{}条用例执行不通过！'.format(case_id))
            final_re = 'Failed'
        write_case(filename,sheetname,case_id + 1,8,final_re)
        print('*'*20)



# 4、将数据写入excel文件，封装成函数
def write_case(filename,sheetname,row,column,result):
    wb = openpyxl.load_workbook(filename)
    sheet= wb[sheetname]
    sheet.cell(row = row,column = column).value = result
    wb.save(filename)

# execute_fun('../test_data/test_case_api.xlsx', 'login')
