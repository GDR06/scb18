# -*- coding: utf-8 -*- 
# @Time : 2020/12/23 17:33 
# @Author : GDR 
# @File : test.py


import requests
import openpyxl
import ast

def api_fun(url,data):           # data是一个dict形式
    header = {'X-Lemonban-Media-Type': 'lemonban.v2', 'Content-Type': 'application/json'}
    result = requests.post(url = url,json = data,headers = header).json()
    return result            # 返回结果是一个dict

def read_case(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row
    case_list = []
    for i in range(2,max_row + 1):
        case_dict = dict(
            id = sheet.cell(row=i, column=1).value,
            url=sheet.cell(row=i, column=5).value,
            data = sheet.cell(row=i, column=6).value,
            expect = sheet.cell(row=i, column=7).value
        )
        case_list.append(case_dict)
    return case_list                        # 返回的是一个list


# 读取到注册的id、url、data、expect
registercase_list = read_case('../test_data/test_case_api.xlsx', 'register')
# print(registercase_list)

list_len = len(registercase_list)


# 获取到注册数据list的data部分
list_data = []
for i in range(0,list_len - 1):
    data= registercase_list[i]['data']         # data的形式是字符串
    list_data.append(data)
print(list_data)                               # list_data是一个由字符串组成的列表

# 发送请求
response_list = []
for i in range(0,list_len - 1):
    url_test = 'http://120.78.128.25:8766/futureloan/member/register'
    res_dict = api_fun(url_test,ast.literal_eval(list_data[i]))
    response_list.append(res_dict)
# print(response_list)                      # 返回的响应数据放在一个list里面

print(response_list[2])
print(registercase_list[2]['expect'])
print(response_list[2]['code'] == ast.literal_eval(registercase_list[2]['expect'])['code'])